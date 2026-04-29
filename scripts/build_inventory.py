#!/usr/bin/env python3
"""
build_inventory.py — Cytomove validation set inventory builder.

Walks the `wound healing/` archive, derives metadata from path structure,
reads image dimensions, parses pre-existing ImageJ ground-truth xlsx
files, emits CSV + summary markdown.

Requires: Pillow and openpyxl.
"""

import csv
import os
import sys
from collections import Counter
from pathlib import Path

from PIL import Image
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
ARCHIVE_ROOT = PROJECT_ROOT / "wound healing"
OUTPUT_CSV = PROJECT_ROOT / "docs" / "validation-inventory.csv"
OUTPUT_MD = PROJECT_ROOT / "docs" / "validation-inventory-summary.md"

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".tif", ".tiff"}

CONDITION_MAP = {
    "K":   ("Control", "untreated"),
    "8F":  ("FDI-6 8 uM", "FDI-6"),
    "64F": ("FDI-6 64 uM", "FDI-6"),
    "CIS": ("Cisplatin", "cisplatin"),
    "CİS": ("Cisplatin", "cisplatin"),
    "LUT": ("Luteolin", "luteolin"),
    "LC":  ("Luteolin + Cisplatin", "combination"),
    "FDI": ("FDI-6", "FDI-6"),
}
CELL_LINE_MAP = {"H": "HUVEC", "M": "MDA-MB-231"}


def derive_metadata(rel_path: Path) -> dict:
    parts = list(rel_path.parts)
    folder = rel_path.parent.name
    md = {
        "cell_line": "", "cell_line_code": "",
        "condition": "", "condition_code": "", "treatment_class": "",
        "time_point_h": "", "campaign": "", "campaign_year": "",
        "in_combine_set": False,
    }
    top = parts[0] if parts else ""

    if top.lower() == "fdi":
        md["campaign"] = "fdi-huvec-2021"
        md["campaign_year"] = "2021"
        md["cell_line"] = "HUVEC"
        md["cell_line_code"] = "H"
        f = folder.lower()
        if "0h" in f:
            md["time_point_h"] = "0"
        elif "48h" in f:
            md["time_point_h"] = "48"
        elif "24h" in f or "24 h" in f:
            md["time_point_h"] = "24"
        if "kontrol" in f or "control" in f:
            md["condition_code"] = "K"
            md["condition"] = "Control"
            md["treatment_class"] = "untreated"
        else:
            md["condition_code"] = "FDI"
            md["condition"] = "FDI-6"
            md["treatment_class"] = "FDI-6"
        return md

    if top == "29.06.22":
        md["campaign"] = "29.06.22"
        md["campaign_year"] = "2022"

        if "COMBİNE" in parts or "COMBINE" in parts:
            md["in_combine_set"] = True
            tokens = folder.split("-")
            if len(tokens) >= 2 and tokens[0] in ("H", "M"):
                md["cell_line_code"] = tokens[0]
                md["cell_line"] = CELL_LINE_MAP.get(tokens[0], "")
                cc = tokens[1]
                cmap = CONDITION_MAP.get(cc, (cc, ""))
                md["condition_code"] = cc
                md["condition"] = cmap[0]
                md["treatment_class"] = cmap[1]
            md["time_point_h"] = "multi"
            return md

        if "yara iyileşme deneyi" in parts:
            f = folder.lower().replace("h", "").strip()
            md["time_point_h"] = f if f.isdigit() else folder
            md["cell_line"] = "unresolved"
            md["condition_code"] = "unresolved"
            md["condition"] = "timecourse-mixed"
            return md

        # Baseline 0H folders: H-K-0H, M-LC-0H, etc.
        tokens = folder.split("-")
        if len(tokens) == 3 and tokens[0] in ("H", "M") and tokens[2] == "0H":
            md["cell_line_code"] = tokens[0]
            md["cell_line"] = CELL_LINE_MAP.get(tokens[0], "")
            cc = tokens[1]
            cmap = CONDITION_MAP.get(cc, (cc, ""))
            md["condition_code"] = cc
            md["condition"] = cmap[0]
            md["treatment_class"] = cmap[1]
            md["time_point_h"] = "0"
            return md

    return md


def get_image_dims(p: Path):
    try:
        with Image.open(p) as im:
            return im.size
    except Exception:
        return (None, None)


def parse_combine_xlsx(combine_dir: Path):
    out = []
    if not combine_dir.exists():
        return out
    for xlsx in combine_dir.rglob("*.xlsx"):
        if xlsx.stem.startswith("~"):
            continue
        folder = xlsx.parent.name
        tokens = folder.split("-")
        cell_code = tokens[0] if tokens and tokens[0] in ("H", "M") else ""
        cond_code = tokens[1] if len(tokens) >= 2 and tokens[0] in ("H", "M") else ""
        try:
            wb = openpyxl.load_workbook(xlsx, data_only=True)
        except Exception as e:
            print(f"   xlsx read failed ({xlsx}): {e}", file=sys.stderr)
            continue
        for sh in wb.sheetnames:
            ws = wb[sh]
            rows = list(ws.iter_rows(values_only=True))
            if not rows or not rows[0]:
                continue
            header = rows[0]
            idx = {"area": None, "label": None}
            for i, c in enumerate(header):
                cs = str(c).lower() if c else ""
                if "area pixels" in cs and idx["area"] is None:
                    idx["area"] = i
                elif cs == "area" and idx["area"] is None:
                    idx["area"] = i
                if "label" in cs and idx["label"] is None:
                    idx["label"] = i
            time_seq = ["0", "24", "48"]
            for ti in range(min(3, len(rows) - 1)):
                row = rows[ti + 1]
                if not row:
                    continue
                area = row[idx["area"]] if idx["area"] is not None and idx["area"] < len(row) else None
                if not isinstance(area, (int, float)):
                    continue
                label = row[idx["label"]] if idx["label"] is not None and idx["label"] < len(row) else None
                out.append({
                    "source_xlsx": str(xlsx.relative_to(combine_dir.parent.parent)),
                    "cell_line_code": cell_code,
                    "condition_code": cond_code,
                    "time_point_h": time_seq[ti],
                    "label_in_xlsx": label or "",
                    "area_pixels2": float(area),
                })
    return out


def main():
    if not ARCHIVE_ROOT.exists():
        print(f"ERROR: archive not found at {ARCHIVE_ROOT}", file=sys.stderr)
        sys.exit(1)
    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)

    measurements = parse_combine_xlsx(ARCHIVE_ROOT / "29.06.22" / "COMBİNE")
    print(f"[xlsx] Parsed {len(measurements)} measurements from COMBİNE/*.xlsx")

    all_images = []
    for r, _, files in os.walk(ARCHIVE_ROOT):
        for fn in files:
            if Path(fn).suffix.lower() in IMAGE_EXTENSIONS:
                full = Path(r) / fn
                rel = full.relative_to(ARCHIVE_ROOT)
                all_images.append((full, rel))
    print(f"[scan] Found {len(all_images)} images")

    rows = []
    for full, rel in all_images:
        md = derive_metadata(rel)
        size_bytes = full.stat().st_size
        w, h = get_image_dims(full)
        gt = ""
        if md.get("in_combine_set"):
            gt = f"condition-level (xlsx for {md['cell_line_code']}-{md['condition_code']})"
        rows.append({
            "filename": rel.name,
            "relative_path": str(rel).replace("\\", "/"),
            "campaign": md["campaign"],
            "campaign_year": md["campaign_year"],
            "cell_line": md["cell_line"],
            "cell_line_code": md["cell_line_code"],
            "condition": md["condition"],
            "condition_code": md["condition_code"],
            "treatment_class": md["treatment_class"],
            "time_point_h": md["time_point_h"],
            "magnification": "10x",
            "microscope_brand": "Olympus",
            "imaging_modality": "brightfield-via-eyepiece",
            "image_width_px": w if w else "",
            "image_height_px": h if h else "",
            "file_size_bytes": size_bytes,
            "in_combine_set": md["in_combine_set"],
            "ground_truth_xlsx_link": gt,
        })

    rows.sort(key=lambda r: (
        r["campaign"], r["cell_line_code"], r["condition_code"],
        str(r["time_point_h"]), r["filename"],
    ))
    for i, r in enumerate(rows, 1):
        r["image_id"] = f"cytv-{i:04d}"

    fieldnames = [
        "image_id", "filename", "relative_path",
        "campaign", "campaign_year",
        "cell_line", "cell_line_code",
        "condition", "condition_code", "treatment_class",
        "time_point_h", "magnification", "microscope_brand", "imaging_modality",
        "image_width_px", "image_height_px", "file_size_bytes",
        "in_combine_set", "ground_truth_xlsx_link",
    ]
    with OUTPUT_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)
    print(f"[csv ] Wrote {OUTPUT_CSV.name} ({len(rows)} rows)")

    by_camp = Counter(r["campaign"] for r in rows)
    by_cell = Counter(r["cell_line"] or "(unresolved)" for r in rows)
    by_cond = Counter(r["condition"] or "(unresolved)" for r in rows)
    by_time = Counter(str(r["time_point_h"]) or "(unresolved)" for r in rows)
    by_dim = Counter(f"{r['image_width_px']}x{r['image_height_px']}" for r in rows)
    n_combine = sum(1 for r in rows if r["in_combine_set"])
    n_unresolved = sum(1 for r in rows if r["condition_code"] in ("", "unresolved"))
    n_labeled = sum(1 for m in measurements if m["label_in_xlsx"])

    crosstab = Counter()
    for r in rows:
        if r["cell_line_code"] in ("H", "M") and r["condition_code"]:
            crosstab[(r["cell_line_code"], r["condition_code"], str(r["time_point_h"]))] += 1

    L = []
    L.append("# Cytomove Validation Inventory — Summary\n")
    L.append(f"**Archive root:** `wound healing/`")
    L.append(f"**Total images:** {len(rows)}")
    L.append("**Generated by:** `scripts/build_inventory.py`\n")

    L.append("## By campaign\n")
    L.append("| Campaign | n |")
    L.append("|---|---|")
    for k, v in sorted(by_camp.items(), key=lambda x: -x[1]):
        L.append(f"| {k} | {v} |")
    L.append("")

    L.append("## By cell line\n")
    L.append("| Cell line | n |")
    L.append("|---|---|")
    for k, v in sorted(by_cell.items(), key=lambda x: -x[1]):
        L.append(f"| {k} | {v} |")
    L.append("")

    L.append("## By condition\n")
    L.append("| Condition | n |")
    L.append("|---|---|")
    for k, v in sorted(by_cond.items(), key=lambda x: -x[1]):
        L.append(f"| {k} | {v} |")
    L.append("")

    L.append("## By time point\n")
    L.append("| Time (h) | n |")
    L.append("|---|---|")
    for k, v in sorted(by_time.items()):
        L.append(f"| {k} | {v} |")
    L.append("")

    L.append("## By image dimensions\n")
    L.append("| WxH (px) | n |")
    L.append("|---|---|")
    for k, v in sorted(by_dim.items(), key=lambda x: -x[1]):
        L.append(f"| {k} | {v} |")
    L.append("")

    L.append("## Tier 1 diversity matrix\n")
    L.append("Counts per (cell line x condition x timepoint).\n")
    L.append("| Cell | Cond | t=0h | t=24h | t=48h | t=multi |")
    L.append("|---|---|---|---|---|---|")
    cells = sorted({k[0] for k in crosstab})
    conds = sorted({k[1] for k in crosstab})
    for cl in cells:
        for cd in conds:
            n0 = crosstab.get((cl, cd, "0"), 0)
            n24 = crosstab.get((cl, cd, "24"), 0)
            n48 = crosstab.get((cl, cd, "48"), 0)
            nm = crosstab.get((cl, cd, "multi"), 0)
            if n0 + n24 + n48 + nm == 0:
                continue
            L.append(f"| {CELL_LINE_MAP.get(cl, cl)} | {cd} | {n0} | {n24} | {n48} | {nm} |")
    L.append("")

    L.append("## Pre-existing ImageJ ground-truth measurements\n")
    L.append(f"Parsed {len(measurements)} measurement rows from `COMBİNE/*.xlsx`.\n")
    L.append("| Source xlsx | Cell | Cond | t (h) | Area (px^2) | Label in xlsx |")
    L.append("|---|---|---|---|---|---|")
    for m in measurements:
        L.append(
            f"| {m['source_xlsx']} | {m['cell_line_code']} | {m['condition_code']} | "
            f"{m['time_point_h']} | {int(m['area_pixels2'])} | {m['label_in_xlsx']} |"
        )
    L.append("")

    L.append("## Coverage assessment\n")
    L.append(f"- COMBINE multi-timepoint subset: **{n_combine}** images of {len(rows)}")
    L.append(f"- Unresolved cell line / condition (timecourse pool): **{n_unresolved}** images")
    L.append(f"- Pre-existing per-image ImageJ measurements with explicit Label: **{n_labeled}** of {len(measurements)} xlsx rows")
    L.append("")
    L.append("**Caveat.** xlsx files are condition-level summaries (3 measurement rows per cell x condition, covering 0/24/48 h, with at most one row carrying an explicit image filename Label). They serve as a calibration check for aggregate trends, but per-image ground truth for Pearson r > 0.9 evaluation must still be (re-)measured in ImageJ on a stratified subset. Recommended target: n = 50-80 images stratified across cell x condition x time.")
    L.append("")

    L.append("## Diversity matrix coverage (Tier 1 only)\n")
    L.append("| Axis | Tier 1 coverage | Stretch source |")
    L.append("|---|---|---|")
    L.append("| Cell types | 2 (HUVEC, MDA-MB-231) | + 1 from public dataset |")
    L.append("| Conditions | 6 (Control, FDI-6 x2, CIS, LUT, LC) | strong |")
    L.append("| Time points | 0, 24, 48 h | mid-closure (6-12 h) optional |")
    L.append("| Magnification | 10x only | + 4x or 20x from public |")
    L.append("| Microscope | Olympus brand, single | + dedicated camera optical system |")
    L.append("| Imaging modality | brightfield via smartphone-eyepiece | + phase-contrast formal capture |")
    L.append("")

    OUTPUT_MD.write_text("\n".join(L), encoding="utf-8")
    print(f"[md  ] Wrote {OUTPUT_MD.name}")


if __name__ == "__main__":
    main()
