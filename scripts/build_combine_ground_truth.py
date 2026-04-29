"""Extract COMBINE ImageJ/MRI Wound Healing Tool measurements.

The COMBINE Excel files contain useful pre-existing area/width/closure values,
but not every row is a one-to-one image measurement. This script makes that
explicit by exporting mapping confidence and source kind for each row.
"""

from __future__ import annotations

import csv
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
ARCHIVE_ROOT = ROOT / "wound healing"
COMBINE_ROOT = ARCHIVE_ROOT / "29.06.22" / "COMBİNE"
INVENTORY = ROOT / "docs" / "validation-inventory.csv"
OUT_CSV = ROOT / "docs" / "combine-ground-truth.csv"
OUT_MD = ROOT / "docs" / "combine-ground-truth-summary.md"

CELL_LINE_MAP = {"H": "HUVEC", "M": "MDA-MB-231"}
CONDITION_MAP = {
    "K": "Control",
    "8F": "FDI-6 8 uM",
    "64F": "FDI-6 64 uM",
}
PREFIX_MAP = {
    ("H", "K"): "HK",
    ("H", "8F"): "H8F",
    ("M", "K"): "MK",
    ("M", "8F"): "M8F",
}
TIME_SEQUENCE = ["0", "24", "48"]


def as_float(value: Any) -> float | str:
    return float(value) if isinstance(value, (int, float)) else ""


def as_intish(value: Any) -> str:
    if not isinstance(value, (int, float)):
        return ""
    return str(int(value)) if float(value).is_integer() else str(float(value))


def folder_meta(folder: str) -> tuple[str, str, str, str]:
    tokens = folder.split("-")
    cell_code = tokens[0] if tokens and tokens[0] in CELL_LINE_MAP else ""
    condition_code = tokens[1] if len(tokens) >= 2 else ""
    return (
        cell_code,
        CELL_LINE_MAP.get(cell_code, ""),
        condition_code,
        CONDITION_MAP.get(condition_code, condition_code),
    )


def load_inventory_index() -> dict[tuple[str, str], dict[str, str]]:
    index: dict[tuple[str, str], dict[str, str]] = {}
    if not INVENTORY.exists():
        return index
    with INVENTORY.open("r", encoding="utf-8", newline="") as fh:
        for row in csv.DictReader(fh):
            index[(row["relative_path"], row["filename"].lower())] = row
    return index


def find_image(folder: Path, filename: str, inventory: dict[tuple[str, str], dict[str, str]]) -> tuple[str, str]:
    rel = (Path("29.06.22") / "COMBİNE" / folder.name / filename).as_posix()
    row = inventory.get((rel, filename.lower()))
    if row:
        return row["image_id"], rel
    if (folder / filename).exists():
        return "", rel
    return "", ""


def inferred_filename(cell_code: str, condition_code: str, time_point: str) -> str:
    prefix = PREFIX_MAP.get((cell_code, condition_code), "")
    return f"{prefix}-{time_point}H.jpg" if prefix else ""


def parse_main_table(path: Path, inventory: dict[tuple[str, str], dict[str, str]]) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    folder = path.parent
    cell_code, cell_line, condition_code, condition = folder_meta(folder.name)
    out: list[dict[str, Any]] = []

    # Standard ImageJ output: first rows after header are 0/24/48.
    # H8F is a replicate-block worksheet, so it is handled only through ORT aggregate rows below.
    if path.name == "H8F.xlsx":
        return []
    is_h8f_layout = False
    for idx, time_point in enumerate(TIME_SEQUENCE, start=1):
        if idx >= len(rows):
            break
        row = rows[idx]
        area = row[2] if len(row) > 2 else None
        if not isinstance(area, (int, float)):
            continue
        label = ""
        area_percent = ""
        width = ""
        sd = ""
        closure = ""
        if is_h8f_layout:
            closure = as_float(row[3] if len(row) > 3 else "")
            area_percent = as_float(row[4] if len(row) > 4 else "")
            width = as_float(row[5] if len(row) > 5 else "")
            sd = as_float(row[6] if len(row) > 6 else "")
        else:
            label = row[1] or "" if len(row) > 1 else ""
            area_percent = as_float(row[3] if len(row) > 3 else "")
            width = as_float(row[4] if len(row) > 4 else "")
            sd = as_float(row[5] if len(row) > 5 else "")
            # Closure is usually column I/O depending on sheet, but the first
            # table's closure is also area-derived; read the nearest populated
            # closure field for the current row when present.
            for col in (6, 8, 14):
                if len(row) > col and isinstance(row[col], (int, float)):
                    closure = float(row[col])
                    break

        filename = str(label).strip() if label else inferred_filename(cell_code, condition_code, time_point)
        image_id, rel = find_image(folder, filename, inventory) if filename else ("", "")
        confidence = "explicit_label" if label and rel else "inferred_from_folder_timepoint" if rel else "condition_timepoint_only"
        out.append(
            {
                "source_kind": "main_table",
                "source_xlsx": path.relative_to(ARCHIVE_ROOT).as_posix(),
                "cell_line_code": cell_code,
                "cell_line": cell_line,
                "condition_code": condition_code,
                "condition": condition,
                "time_point_h": time_point,
                "label_in_xlsx": label,
                "matched_filename": filename,
                "matched_image_id": image_id,
                "matched_relative_path": rel,
                "mapping_confidence": confidence,
                "area_pixels2": as_intish(area),
                "area_percent": area_percent,
                "width_pixels": width,
                "standard_deviation_pixels": sd,
                "wound_closure_percent": closure,
                "replicate_count": "",
                "notes": "Primary 0/24/48 row from Excel top table.",
            }
        )
    return out


def parse_replicate_average_blocks(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    folder = path.parent
    cell_code, cell_line, condition_code, condition = folder_meta(folder.name)
    out: list[dict[str, Any]] = []

    ort_rows = []
    for row_number, row in enumerate(rows, start=1):
        cells = list(row)
        if "ORT" not in cells:
            continue
        # Ignore the compact H8F top-table averages? No: they are useful as
        # aggregate calibration rows, just not one-to-one image labels.
        area = ""
        closure = ""
        area_percent = ""
        try:
            ort_idx = cells.index("ORT")
        except ValueError:
            continue
        for value in cells[ort_idx + 1 :]:
            if isinstance(value, (int, float)):
                area = as_intish(value)
                break
        # Closure/area percent is valid only when it is the immediate cell after the area.
        # Some sheets place unrelated summary tables farther to the right.
        immediate_after_area_idx = ort_idx + 2
        if len(cells) > immediate_after_area_idx and isinstance(cells[immediate_after_area_idx], (int, float)):
            closure = float(cells[immediate_after_area_idx])
            area_percent = closure
        if area:
            ort_rows.append((row_number, area, area_percent, closure))

    for idx, (row_number, area, area_percent, closure) in enumerate(ort_rows[:3]):
        time_point = TIME_SEQUENCE[idx]
        out.append(
            {
                "source_kind": "replicate_average",
                "source_xlsx": path.relative_to(ARCHIVE_ROOT).as_posix(),
                "cell_line_code": cell_code,
                "cell_line": cell_line,
                "condition_code": condition_code,
                "condition": condition,
                "time_point_h": time_point,
                "label_in_xlsx": "",
                "matched_filename": "",
                "matched_image_id": "",
                "matched_relative_path": "",
                "mapping_confidence": "aggregate_no_single_image",
                "area_pixels2": area,
                "area_percent": area_percent,
                "width_pixels": "",
                "standard_deviation_pixels": "",
                "wound_closure_percent": closure,
                "replicate_count": "2",
                "notes": f"ORT aggregate row {row_number}; use for trend calibration, not one-to-one image validation.",
            }
        )
    return out


def write_summary(rows: list[dict[str, Any]]) -> None:
    by_kind = Counter(row["source_kind"] for row in rows)
    by_conf = Counter(row["mapping_confidence"] for row in rows)
    by_group = Counter((row["cell_line"], row["condition"]) for row in rows if row["source_kind"] == "main_table")

    lines = [
        "# Cytomove COMBINE Ground Truth — Summary",
        "",
        "**Generated by:** `scripts/build_combine_ground_truth.py`  ",
        "**Input:** `wound healing/29.06.22/COMBİNE/*.xlsx`  ",
        "**Output:** `docs/combine-ground-truth.csv`",
        "",
        f"**Total exported rows:** {len(rows)}",
        "",
        "## By Source Kind",
        "",
        "| Source kind | n |",
        "|---|---:|",
    ]
    for key, value in sorted(by_kind.items()):
        lines.append(f"| {key} | {value} |")
    lines.extend(["", "## By Mapping Confidence", "", "| Mapping confidence | n |", "|---|---:|"])
    for key, value in sorted(by_conf.items()):
        lines.append(f"| {key} | {value} |")
    lines.extend(["", "## Main Table Groups", "", "| Cell line | Condition | n |", "|---|---|---:|"])
    for (cell, condition), value in sorted(by_group.items()):
        lines.append(f"| {cell} | {condition} | {value} |")
    lines.extend(
        [
            "",
            "## Interpretation",
            "",
            "- Use `area_pixels2`, `area_percent`, and `wound_closure_percent` as the first validation/calibration target for the COMBINE subset.",
            "- Use `width_pixels` and `standard_deviation_pixels` as the second metric family after the area-based segmentation path is stable.",
            "- Rows with `mapping_confidence=explicit_label` are safest for one-to-one image validation.",
            "- Rows with `mapping_confidence=inferred_from_folder_timepoint` are useful, but should be visually checked against the corresponding 0h/24h/48h image before being treated as final per-image ground truth.",
            "- Rows with `source_kind=replicate_average` are aggregate calibration evidence, not single-image labels.",
        ]
    )
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    inventory = load_inventory_index()
    rows: list[dict[str, Any]] = []
    for path in sorted(COMBINE_ROOT.rglob("*.xlsx")):
        if path.name.startswith("~"):
            continue
        rows.extend(parse_main_table(path, inventory))
        rows.extend(parse_replicate_average_blocks(path))

    fieldnames = [
        "source_kind",
        "source_xlsx",
        "cell_line_code",
        "cell_line",
        "condition_code",
        "condition",
        "time_point_h",
        "label_in_xlsx",
        "matched_filename",
        "matched_image_id",
        "matched_relative_path",
        "mapping_confidence",
        "area_pixels2",
        "area_percent",
        "width_pixels",
        "standard_deviation_pixels",
        "wound_closure_percent",
        "replicate_count",
        "notes",
    ]
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    write_summary(rows)
    print(f"Exported {len(rows)} rows")
    print(f"Wrote {OUT_CSV.relative_to(ROOT).as_posix()}")
    print(f"Wrote {OUT_MD.relative_to(ROOT).as_posix()}")


if __name__ == "__main__":
    main()